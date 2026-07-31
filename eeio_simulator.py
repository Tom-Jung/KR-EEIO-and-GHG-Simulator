"""
eeio_simulator.py
─────────────────────────────────────────────────────────────────────────────
한국 EEIO(환경산업연관표) 기반 Scope 1·2·3 탄소배출 시뮬레이터 — 핵심 모듈

외부에서 호출하는 함수:
    load_eeio_data(src_path)          → EEIO 원본 데이터 로딩
    run_simulation(data, params)      → 시뮬레이션 실행 및 결과 반환
    save_excel(result, output_dir)    → 결과 엑셀 저장

노트북에서는 이 모듈을 import해서 두 줄이면 실행됩니다.
"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime
import glob
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────────────────────────────────────
VA_NAMES = ["피용자보수", "영업잉여", "고정자본소모"]
TAX_NAME = "생산세(보조금공제)"
TOL = 1e-6

FD_NAMES = {
    "9111": "민간소비지출", "9112": "정부소비지출",
    "9121": "민간고정자본형성", "9122": "정부고정자본형성",
    "9131": "재고증감", "9132": "귀중품 순취득", "9140": "수출",
}

# 산업코드 → 한글명 (원본 파일에서 자동 읽지만, 미리 비워 둠)
_INDUSTRY_NAMES_FALLBACK: dict[str, str] = {}


# ──────────────────────────────────────────────────────────────────────────────
# 예외
# ──────────────────────────────────────────────────────────────────────────────
class 처리불가능(Exception):
    """입력 파라미터로는 분석표를 구성할 수 없을 때."""


# ══════════════════════════════════════════════════════════════════════════════
# 1. 원본 파일 탐지 & 로딩
# ══════════════════════════════════════════════════════════════════════════════

def find_simulator_file(search_dir: str | Path | None = None) -> Path:
    """같은 폴더에 있는 2023_Simulator*.xlsx 파일을 자동 탐지한다."""
    dirs = []
    if search_dir:
        dirs.append(Path(search_dir))

    # IPython(Jupyter) 환경이면 노트북 위치도 추가
    try:
        import IPython
        ip = IPython.get_ipython()
        if ip:
            for attr in ("__vsc_ipynb_file__", "_file_"):
                nb_file = getattr(ip, attr, None)
                if nb_file:
                    dirs.append(Path(nb_file).parent)
                    break
    except Exception:
        pass

    dirs.append(Path.cwd())

    patterns = ["*Simulator*.xlsx"]
    for d in dirs:
        for pat in patterns:
            hits = sorted(d.glob(pat))
            hits = [p for p in hits if p.parent.name != "output"]
            if hits:
                return hits[0]

    raise FileNotFoundError(
        "이 노트북과 같은 폴더에 '2023_Simulator…' 로 시작하는 시뮬레이터 템플릿 "
        ".xlsx 파일(원가전부 / 일부원가 / 매출액만 버전 중 하나)을 두어야 합니다."
    )


# ──────────────────────────────────────────────────────────────────────────────
# 내부: 시트 구조 파싱
# ──────────────────────────────────────────────────────────────────────────────

def _find_data_start_row(ws, max_scan=20):
    for r in range(1, max_scan + 1):
        if ws.cell(row=r, column=1).value == "A":
            return r
    raise ValueError("산업분류 시작 행을 찾을 수 없습니다")


def _find_data_start_col(ws, header_row, max_scan=20):
    for c in range(1, max_scan + 1):
        if ws.cell(row=header_row, column=c).value == "A":
            return c
    raise ValueError("산업분류 시작 열을 찾을 수 없습니다")


def _parse_io_structure(ws, has_value_added=True, has_ghg=False):
    data_start_row = _find_data_start_row(ws)
    header_code_row = data_start_row - 2
    header_name_row = data_start_row - 1

    row_order, row_index = [], {}
    r = data_start_row
    while True:
        code = ws.cell(row=r, column=1).value
        if code is None:
            raise ValueError(f"{r}행: 산업 코드가 비었습니다")
        if code == "9590":
            inter_total_row = r
            break
        row_order.append(code)
        row_index[code] = r
        r += 1

    va_rows, va_total_row, total_input_row, ghg_row = {}, None, None, None
    if has_value_added:
        r = inter_total_row + 1
        while True:
            code = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            if code == "9690":
                va_total_row = r
                r += 1
                if ws.cell(row=r, column=1).value == "9790":
                    total_input_row = r
                break
            va_rows[name] = r
            r += 1
        if has_ghg and ws.cell(row=total_input_row + 1, column=1).value == "GHG":
            ghg_row = total_input_row + 1

    data_start_col = _find_data_start_col(ws, header_code_row)
    col_order, col_index = [], {}
    c = data_start_col
    while True:
        code = ws.cell(row=header_code_row, column=c).value
        if code is None:
            raise ValueError(f"{get_column_letter(c)}열: 산업 코드가 비었습니다")
        if code == "9090":
            inter_demand_col = c
            break
        col_order.append(code)
        col_index[code] = c
        c += 1

    fd_cols, fd_total_col, total_demand_col = {}, None, None
    c = inter_demand_col + 1
    while True:
        code = ws.cell(row=header_code_row, column=c).value
        if code == "9190":
            fd_total_col = c
        elif code == "9290":
            total_demand_col = c
            break
        else:
            fd_cols[code] = c
        c += 1

    assert row_order == col_order, "행/열 산업분류 순서가 일치하지 않습니다"

    return dict(
        data_start_row=data_start_row, header_code_row=header_code_row,
        header_name_row=header_name_row, row_order=row_order, row_index=row_index,
        inter_total_row=inter_total_row, va_rows=va_rows, va_total_row=va_total_row,
        total_input_row=total_input_row, ghg_row=ghg_row,
        data_start_col=data_start_col, col_order=col_order, col_index=col_index,
        inter_demand_col=inter_demand_col, fd_cols=fd_cols,
        fd_total_col=fd_total_col, total_demand_col=total_demand_col,
    )


def _parse_coef_structure(ws):
    data_start_row = _find_data_start_row(ws)
    header_code_row = data_start_row - 2
    row_order, row_index = [], {}
    r = data_start_row
    while True:
        code = ws.cell(row=r, column=1).value
        if code == "9590":
            inter_total_row = r
            break
        row_order.append(code)
        row_index[code] = r
        r += 1
    va_rows = {}
    r = inter_total_row + 1
    while True:
        code = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        if code == "9690" or code is None:
            break
        va_rows[name] = r
        r += 1
    data_start_col = _find_data_start_col(ws, header_code_row)
    col_order, col_index = [], {}
    c = data_start_col
    while True:
        code = ws.cell(row=header_code_row, column=c).value
        if code is None or code == "9090":
            break
        col_order.append(code)
        col_index[code] = c
        c += 1
    return dict(row_index=row_index, col_index=col_index, va_rows=va_rows)


def _get_industry_names(ws, struct):
    return {code: ws.cell(row=r, column=2).value for code, r in struct["row_index"].items()}


def _extract_arrays(ws, struct):
    codes = struct["row_order"]
    n = len(codes)
    fd_codes = sorted(struct["fd_cols"])
    n_fd = len(fd_codes)

    mat = np.zeros((n, n))
    fd = np.zeros((n, n_fd))
    for i, ci in enumerate(codes):
        ri = struct["row_index"][ci]
        for j, cj in enumerate(codes):
            v = ws.cell(row=ri, column=struct["col_index"][cj]).value
            mat[i, j] = v if v is not None else 0.0
        for k, fc in enumerate(fd_codes):
            v = ws.cell(row=ri, column=struct["fd_cols"][fc]).value
            fd[i, k] = v if v is not None else 0.0

    out = dict(codes=codes, fd_codes=fd_codes, mat=mat, fd=fd)

    if struct.get("va_rows"):
        va = np.zeros((3, n))
        tax = np.zeros(n)
        for j, cj in enumerate(codes):
            col = struct["col_index"][cj]
            for v_i, name in enumerate(VA_NAMES):
                vv = ws.cell(row=struct["va_rows"][name], column=col).value
                va[v_i, j] = vv if vv is not None else 0.0
            tv = ws.cell(row=struct["va_rows"][TAX_NAME], column=col).value
            tax[j] = tv if tv is not None else 0.0
        out["va"] = va
        out["tax"] = tax

        total_output = np.zeros(n)
        for j, cj in enumerate(codes):
            tv = ws.cell(row=struct["total_input_row"], column=struct["col_index"][cj]).value
            total_output[j] = tv if tv is not None else 0.0
        out["total_output"] = total_output

    if struct.get("ghg_row"):
        ghg = np.zeros(n)
        for j, cj in enumerate(codes):
            gv = ws.cell(row=struct["ghg_row"], column=struct["col_index"][cj]).value
            ghg[j] = gv if gv is not None else 0.0
        out["ghg"] = ghg

    return out


def _extract_coef_arrays(ws, coef_struct, codes):
    n = len(codes)
    mat = np.zeros((n, n))
    va = np.zeros((3, n))
    for j, cj in enumerate(codes):
        cidx = coef_struct["col_index"][cj]
        for i, ci in enumerate(codes):
            v = ws.cell(row=coef_struct["row_index"][ci], column=cidx).value
            mat[i, j] = v if v is not None else 0.0
        for v_i, name in enumerate(VA_NAMES):
            v = ws.cell(row=coef_struct["va_rows"][name], column=cidx).value
            va[v_i, j] = v if v is not None else 0.0
    return mat, va


# ══════════════════════════════════════════════════════════════════════════════
# 2. 공개 API: 데이터 로딩
# ══════════════════════════════════════════════════════════════════════════════

def load_eeio_data(src_path: str | Path | None = None) -> dict:
    """
    시뮬레이터 템플릿 xlsx 파일에서 EEIO 원본 데이터를 로딩한다.

    Parameters
    ----------
    src_path : str | Path | None
        xlsx 파일 경로. None 이면 현재 폴더에서 자동 탐지.

    Returns
    -------
    dict  {
        "src_path"       : Path,
        "output_dir"     : Path,
        "codes"          : list[str],          # 33개 산업 코드
        "industry_names" : dict[str, str],
        "A_data"         : dict,               # 총거래표 배열
        "D_data"         : dict,               # 국산거래표 배열
        "coef_mat"       : np.ndarray,         # 총투입계수
    }
    """
    src = Path(src_path) if src_path else find_simulator_file()
    output_dir = src.parent / "output"
    output_dir.mkdir(exist_ok=True)

    print(f"📂 원본 파일: {src.name}")

    wb = openpyxl.load_workbook(src, data_only=True)

    ws_A = wb["A표_총거래표(생산자)_원본"]
    A_struct = _parse_io_structure(ws_A, has_value_added=True, has_ghg=True)
    industry_names = _get_industry_names(ws_A, A_struct)
    A_data = _extract_arrays(ws_A, A_struct)

    ws_D = wb["국산거래표_원본"]
    D_struct = _parse_io_structure(ws_D, has_value_added=False)
    D_data = _extract_arrays(ws_D, D_struct)

    ws_coef = wb["총투입계수(A)"]
    coef_struct = _parse_coef_structure(ws_coef)
    coef_mat, _ = _extract_coef_arrays(ws_coef, coef_struct, A_data["codes"])

    print(f"✅ 데이터 로딩 완료  (산업 {len(A_data['codes'])}개)")

    return dict(
        src_path=src,
        output_dir=output_dir,
        codes=A_data["codes"],
        industry_names=industry_names,
        A_data=A_data,
        D_data=D_data,
        coef_mat=coef_mat,
    )


# ══════════════════════════════════════════════════════════════════════════════
# 3. 내부: 원가 배분 / 행·열 삽입 / Ad·Lf·M 계산
# ══════════════════════════════════════════════════════════════════════════════

def _allocate_costs(known_intermediate, known_va, codes, revenue, coef_mat, host_idx):
    n = len(codes)
    L = np.zeros(n)
    is_known = np.zeros(n, dtype=bool)
    for i, c in enumerate(codes):
        if c in known_intermediate:
            L[i] = known_intermediate[c]
            is_known[i] = True

    L_va = np.zeros(3)
    for v_i, name in enumerate(VA_NAMES):
        if name in known_va:
            L_va[v_i] = known_va[name]

    known_total = L[is_known].sum() + L_va.sum()
    remaining = revenue - known_total

    if remaining < -TOL:
        raise 처리불가능(
            f"⛔ 처리 불가능: 아는 비용 합({known_total:,.1f})이 매출액({revenue:,.1f})을 초과합니다."
        )

    K_est = coef_mat[:, host_idx] * revenue
    denom = K_est[~is_known].sum()
    if remaining > TOL:
        if denom <= TOL:
            raise 처리불가능(
                "⛔ 처리 불가능: 잔여예산을 배분할 미상(未詳) 중간투입 항목이 없습니다."
            )
        L[~is_known] = K_est[~is_known] / denom * remaining

    return L, L_va


def _four_point_balance(mat, h, company_idx):
    """Pure 4-Point 교차 밸런싱 (음수 셀 제거)."""
    N = mat.shape[0]
    for i in range(N):
        if mat[i, h] < -1e-6:
            deficit = abs(mat[i, h]); mat[i, h] = 0.0
            tr = mat[i, :].copy(); tr[h] = -np.inf; tr[company_idx] = -np.inf
            mc = np.argmax(tr)
            tc = mat[:, h].copy(); tc[i] = -np.inf; tc[company_idx] = -np.inf
            mr = np.argmax(tc)
            mat[i, mc] -= deficit; mat[mr, h] -= deficit; mat[mr, mc] += deficit
    for j in range(N):
        if mat[h, j] < -1e-6:
            deficit = abs(mat[h, j]); mat[h, j] = 0.0
            tr = mat[h, :].copy(); tr[j] = -np.inf; tr[company_idx] = -np.inf
            mc = np.argmax(tr)
            tc = mat[:, j].copy(); tc[h] = -np.inf; tc[company_idx] = -np.inf
            mr = np.argmax(tc)
            mat[h, mc] -= deficit; mat[mr, j] -= deficit; mat[mr, mc] += deficit


def _build_A_new(A_data, host_idx, revenue, L, L_va):
    codes = A_data["codes"]; n = len(codes)
    mat, fd, va, tax = A_data["mat"], A_data["fd"], A_data["va"], A_data["tax"]
    n_fd = fd.shape[1]; h = host_idx; ci = h + 1; N = n + 1

    def m(i): return i if i <= h else i + 1

    T_h = mat[h, :].sum() + fd[h, :].sum()
    scale = revenue / T_h if T_h > 0 else 0.0

    nm = np.zeros((N, N)); nf = np.zeros((N, n_fd))
    nv = np.zeros((3, N)); nt = np.zeros(N)

    for i in range(n):
        if i != h:
            for j in range(n):
                if j != h:
                    nm[m(i), m(j)] = mat[i, j]
            nm[m(i), h] = mat[i, h] - L[i]
            nf[m(i), :] = fd[i, :]
        nm[m(i), ci] = L[i]

    nm[ci, ci] = 0.0; nm[h, ci] = L[h]
    for j in range(n):
        if j != h:
            nm[ci, m(j)] = mat[h, j] * scale
    nm[ci, h] = mat[h, h] * scale
    for k in range(n_fd):
        nf[ci, k] = fd[h, k] * scale

    for j in range(n):
        if j != h:
            nm[h, m(j)] = mat[h, j] - nm[ci, m(j)]
    nm[h, h] = mat[h, h] - nm[h, ci] - nm[ci, h]
    for k in range(n_fd):
        nf[h, k] = fd[h, k] - nf[ci, k]

    for j in range(n):
        nj = m(j)
        if j != h:
            nv[:, nj] = va[:, j]; nt[nj] = tax[j]
        else:
            nv[:, ci] = L_va; nv[:, h] = va[:, j] - L_va
            nt[ci] = 0.0; nt[h] = tax[j]

    _four_point_balance(nm, h, ci)

    for v in range(3):
        if nv[v, h] < -1e-6:
            deficit = abs(nv[v, h]); nv[v, h] = 0.0
            tmp = nv[:, h].copy(); tmp[v] = -np.inf
            mv = np.argmax(tmp); nv[mv, h] -= deficit

    total_out = nm.sum(axis=0) + nv.sum(axis=0) + nt
    new_codes = codes[:ci] + ["분석업체"] + codes[ci:]
    return dict(codes=new_codes, mat=nm, fd=nf, va=nv, tax=nt,
                total_output=total_out, host_idx=h, company_idx=ci)


def _build_D_new(D_data, A_data, A_new, host_idx):
    codes = D_data["codes"]; n = len(codes)
    dmat, dfd = D_data["mat"], D_data["fd"]
    amat = A_data["mat"]; h = host_idx; ci = h + 1; N = n + 1
    full = A_new["mat"]

    def m(i): return i if i <= h else i + 1
    def ratio(a, b): return a / b if abs(b) > 1e-9 else 0.0

    nm = np.zeros((N, N)); nf = np.zeros((N, dfd.shape[1]))

    for i in range(n):
        if i != h:
            for j in range(n):
                if j != h:
                    nm[m(i), m(j)] = dmat[i, j]
            nm[m(i), ci] = ratio(dmat[i, h], amat[i, h]) * full[m(i), ci]
            nm[m(i), h] = dmat[i, h] - nm[m(i), ci]
            nf[m(i), :] = dfd[i, :]

    nm[ci, ci] = 0.0; nm[h, ci] = ratio(dmat[h, h], amat[h, h]) * full[h, ci]
    for j in range(N):
        if j != ci:
            nm[ci, j] = full[ci, j]
    for k in range(dfd.shape[1]):
        nf[ci, k] = A_new["fd"][ci, k]

    for j in range(n):
        if j != h:
            nm[h, m(j)] = dmat[h, j] - nm[ci, m(j)]
    nm[h, h] = dmat[h, h] - nm[h, ci] - nm[ci, h]
    for k in range(dfd.shape[1]):
        nf[h, k] = dfd[h, k] - nf[ci, k]

    _four_point_balance(nm, h, ci)

    new_codes = codes[:ci] + ["분석업체"] + codes[ci:]
    return dict(codes=new_codes, mat=nm, fd=nf)


def _build_ghg_new(A_data, host_idx, revenue):
    codes = A_data["codes"]; ghg = A_data["ghg"]; total = A_data["total_output"]
    n = len(codes); h = host_idx; ci = h + 1; N = n + 1

    def m(i): return i if i <= h else i + 1

    ng = np.zeros(N)
    for j in range(n):
        nj = m(j)
        if j != h:
            ng[nj] = ghg[j]
        else:
            cg = (ghg[j] / total[j] * revenue) if total[j] > 1e-9 else 0.0
            ng[ci] = cg; ng[h] = max(0.0, ghg[j] - cg)
    return ng


def _compute_ad_lf_m(A_new, D_new, new_ghg):
    total_output = A_new["total_output"]; N = len(total_output)
    with np.errstate(divide="ignore", invalid="ignore"):
        Ad = np.where(total_output[None, :] > 1e-9,
                      D_new["mat"] / total_output[None, :], 0.0)
        ghg_coef = np.where(total_output > 1e-9, new_ghg / total_output, 0.0)
    Lf = np.linalg.inv(np.eye(N) - Ad)
    M = ghg_coef[:, None] * Lf
    return Ad, Lf, M, ghg_coef, M.sum(axis=0)


# ══════════════════════════════════════════════════════════════════════════════
# 4. 공개 API: 시뮬레이션 실행
# ══════════════════════════════════════════════════════════════════════════════

def run_simulation(data: dict, params: dict) -> dict:
    """
    EEIO 탄소배출 시뮬레이션을 실행한다.

    Parameters
    ----------
    data : load_eeio_data()가 반환한 dict
    params : {
        "company_name" : str,          # 기업명 (결과 파일명에 사용)
        "company_code" : str,          # 소속 산업 코드 (예: "C05")
        "sales"        : float,        # 매출액 (백만원)
        "cost_ratios"  : dict[str, float],   # 원가 비중 (%, 모르면 0)
        "va_ratios"    : dict[str, float],   # 부가가치 비중 (%, 모르면 0)
    }

    Returns
    -------
    dict  {
        "company_name", "company_code", "industry_name",
        "sales", "case_label",
        "scope1", "scope2", "scope3", "total_emission",
        "scope_df"        : pd.DataFrame,   # Scope 요약
        "industry_emit_df": pd.DataFrame,   # 산업별 배출량 기여
        "A_new", "D_new", "Ad", "Lf", "M", "new_ghg",
        "L", "L_va", "codes_new",
    }
    """
    codes       = data["codes"]
    names       = data["industry_names"]
    A_data      = data["A_data"]
    D_data      = data["D_data"]
    coef_mat    = data["coef_mat"]

    company_name = params["company_name"]
    company_code = params["company_code"]
    sales        = params["sales"]
    cost_ratios  = params.get("cost_ratios", {c: 0 for c in codes})
    va_ratios    = params.get("va_ratios", {n: 0 for n in VA_NAMES})

    # ── 검증 ──
    def _is_num(x): return isinstance(x, (int, float)) and not isinstance(x, bool)

    if not _is_num(sales) or sales <= 0:
        raise 처리불가능(f"⛔ 매출액은 0보다 큰 숫자여야 합니다. (입력값: {sales!r})")
    if company_code not in codes:
        raise 처리불가능(f"⛔ 산업 코드 '{company_code}'를 찾을 수 없습니다. 가능한 코드: {codes}")
    for label, d in [("원가", cost_ratios), ("부가가치", va_ratios)]:
        for k, v in d.items():
            if not _is_num(v):
                raise 처리불가능(f"⛔ {label} '{k}' 값이 숫자가 아닙니다: {v!r}")
            if v < 0:
                raise 처리불가능(f"⛔ {label} '{k}' 값이 음수입니다: {v}")
    unknown = [c for c in cost_ratios if c not in codes]
    if unknown:
        raise 처리불가능(f"⛔ 알 수 없는 원가 코드: {unknown}")

    cost_pct = sum(cost_ratios.values())
    va_pct   = sum(va_ratios.values())
    total_pct = cost_pct + va_pct
    if total_pct > 100.0 + TOL:
        raise 처리불가능(
            f"⛔ 원가({cost_pct}%) + 부가가치({va_pct}%) = {total_pct}% > 100%  "
            "— 매출액보다 큰 비용은 배분할 수 없습니다."
        )

    if total_pct >= 100.0 - TOL:
        case_label = "Case3 (원가전부 아는 경우)"
    elif total_pct > 0:
        case_label = "Case2 (일부원가만 아는 경우)"
    else:
        case_label = "Case1 (매출액만 아는 경우)"

    known_inter = {c: p / 100 * sales for c, p in cost_ratios.items() if p > 0}
    known_va    = {n: p / 100 * sales for n, p in va_ratios.items() if p > 0}

    host_idx = codes.index(company_code)
    L, L_va  = _allocate_costs(known_inter, known_va, codes, sales, coef_mat, host_idx)

    col_in = L.sum() + L_va.sum()
    if abs(col_in - sales) > 1e-3:
        raise 처리불가능(f"⛔ 원가 배분 합({col_in:,.3f}) ≠ 매출액({sales:,.3f})")

    # ── 행/열 삽입 ──
    A_new   = _build_A_new(A_data, host_idx, sales, L, L_va)
    D_new   = _build_D_new(D_data, A_data, A_new, host_idx)
    new_ghg = _build_ghg_new(A_data, host_idx, sales)

    ci = A_new["company_idx"]
    Ad, Lf, M, ghg_coef, total_induced = _compute_ad_lf_m(A_new, D_new, new_ghg)

    elec_orig = codes.index("D")
    elec_new  = elec_orig if elec_orig <= host_idx else elec_orig + 1

    scope1 = float(ghg_coef[ci] * sales)
    scope2 = float(M[elec_new, ci] * sales)
    total_emission = float(total_induced[ci] * sales)
    scope3 = total_emission - scope1 - scope2

    # ── Scope 요약 DataFrame ──
    scope_df = pd.DataFrame({
        "항목": ["Scope1 (직접배출)", "Scope2 (전력 등 간접배출)", "Scope3 (기타 간접배출)"],
        "배출량 (tCO2eq.)": [scope1, scope2, scope3],
        "비중": [
            f"{scope1 / total_emission * 100:.1f}%" if total_emission else "–",
            f"{scope2 / total_emission * 100:.1f}%" if total_emission else "–",
            f"{scope3 / total_emission * 100:.1f}%" if total_emission else "–",
        ],
    })

    # ── 산업별 배출량 기여 DataFrame ──
    new_codes = A_new["codes"]
    m_col = M[:, ci] * sales          # 산업별 유발배출량
    industry_emit = []
    for idx, c in enumerate(new_codes):
        nm = "분석업체" if c == "분석업체" else names.get(c, c)
        industry_emit.append({
            "코드": c if c != "분석업체" else "–",
            "산업명": nm,
            "배출량 (tCO2eq.)": float(m_col[idx]),
            "비중": f"{m_col[idx] / total_emission * 100:.1f}%" if total_emission else "–",
        })
    industry_emit_df = (
        pd.DataFrame(industry_emit)
        .sort_values("배출량 (tCO2eq.)", ascending=False)
        .reset_index(drop=True)
    )

    return dict(
        company_name=company_name,
        company_code=company_code,
        industry_name=names.get(company_code, company_code),
        sales=sales,
        case_label=case_label,
        scope1=scope1, scope2=scope2, scope3=scope3,
        total_emission=total_emission,
        scope_df=scope_df,
        industry_emit_df=industry_emit_df,
        A_new=A_new, D_new=D_new,
        Ad=Ad, Lf=Lf, M=M,
        new_ghg=new_ghg,
        L=L, L_va=L_va,
        codes_new=new_codes,
        industry_names=names,
        fd_codes=A_data["fd_codes"],
    )


# ══════════════════════════════════════════════════════════════════════════════
# 5. 공개 API: 결과 시각화
# ══════════════════════════════════════════════════════════════════════════════

def _html_table(df, col_formats=None, row_colors=None, title=None, center_cols=None,
                total_row=None):
    col_formats = col_formats or {}
    th_base = (
        "background:#2c3e50;color:white;padding:9px 16px;"
        "font-size:13px;white-space:nowrap;"
    )
    td_base = "padding:8px 16px;font-size:13px;border-bottom:1px solid #e0e0e0;"
    tf_base = (
        "padding:8px 16px;font-size:13px;font-weight:bold;"
        "border-top:2px solid #2c3e50;background:#f0f4f8;"
    )

    def _align(col):
        if center_cols is None or col in center_cols:
            return "text-align:center;"
        return "text-align:left;"

    rows_html = []
    for _, row in df.iterrows():
        extra = row_colors(row) if row_colors else ""
        cells = []
        for col in df.columns:
            val = row[col]
            if col in col_formats and isinstance(val, (int, float)):
                val = col_formats[col].format(val)
            cells.append(f"<td style='{td_base}{_align(col)}{extra}'>{val}</td>")
        rows_html.append(f"<tr>{''.join(cells)}</tr>")

    if total_row is not None:
        total_cells = []
        for col in df.columns:
            val = total_row.get(col, "")
            if col in col_formats and isinstance(val, (int, float)):
                val = col_formats[col].format(val)
            total_cells.append(f"<td style='{tf_base}{_align(col)}'>{val}</td>")
        rows_html.append(f"<tr>{''.join(total_cells)}</tr>")

    headers = "".join(
        f"<th style='{th_base}{_align(c)}'>{c}</th>" for c in df.columns
    )
    table = (
        f"<table style='border-collapse:collapse;width:100%;margin-bottom:4px'>"
        f"<thead><tr>{headers}</tr></thead>"
        f"<tbody>{''.join(rows_html)}</tbody>"
        f"</table>"
    )
    if title:
        table = f"<h3 style='margin-bottom:6px'>{title}</h3>" + table
    return table


def display_results(result: dict) -> None:
    """
    run_simulation() 결과를 Jupyter 노트북에 HTML 테이블로 표시한다.

    Parameters
    ----------
    result : run_simulation()이 반환한 dict
    """
    from IPython.display import display, HTML

    # ── 기업 정보 ──
    info_df = pd.DataFrame({
        "항목": ["기업명", "소속 산업", "매출액 (백만원)", "계산 모드"],
        "내용": [
            result["company_name"],
            f"{result['company_code']}  {result['industry_name']}",
            f"{result['sales']:,.0f}",
            result["case_label"],
        ],
    })
    display(HTML(_html_table(info_df, center_cols=["내용"], title="🏢 기업 정보")))

    # ── Scope 요약 ──
    SCOPE_COLORS = {"Scope1": "#e63946", "Scope2": "#e07b39", "Scope3": "#457b9d"}

    def _scope_color(row):
        for k, c in SCOPE_COLORS.items():
            if k in str(row["항목"]):
                return f"background:{c};color:white;font-weight:bold;"
        return ""

    display(HTML(_html_table(
        result["scope_df"],
        col_formats={"배출량 (tCO2eq.)": "{:,.1f}"},
        row_colors=_scope_color,
        center_cols=["배출량 (tCO2eq.)", "비중"],
        title="🌍 탄소배출 결과 — Scope 1 · 2 · 3",
        total_row={
            "항목": "합계 (Scope 1+2+3)",
            "배출량 (tCO2eq.)": result["total_emission"],
            "비중": "100.0%",
        },
    )))

    # ── 산업별 배출 기여 ──
    emit_df = result["industry_emit_df"].copy()
    emit_df.insert(0, "순위", range(1, len(emit_df) + 1))

    def _emit_color(row):
        if row["산업명"] == "분석업체":
            return "background:#e63946;color:white;font-weight:bold;"
        if row["순위"] == 1:
            return "background:#ffd166;font-weight:bold;"
        if row["순위"] <= 3:
            return "background:#fff3cd;"
        return ""

    display(HTML(_html_table(
        emit_df,
        col_formats={"배출량 (tCO2eq.)": "{:,.1f}"},
        row_colors=_emit_color,
        center_cols=["순위", "배출량 (tCO2eq.)", "비중"],
        title="🏭 산업별 배출 기여 (전체 / 분석업체는 직접배출)",
        total_row={
            "순위": "",
            "코드": "",
            "산업명": "합계",
            "배출량 (tCO2eq.)": result["total_emission"],
            "비중": "100.0%",
        },
    )))


# ══════════════════════════════════════════════════════════════════════════════
# 6. 공개 API: 엑셀 저장
# ══════════════════════════════════════════════════════════════════════════════

def save_excel(result: dict, output_dir: str | Path) -> Path:
    """시뮬레이션 결과를 엑셀 파일로 저장한다."""
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    A_new      = result["A_new"]
    D_new      = result["D_new"]
    Ad         = result["Ad"]
    Lf         = result["Lf"]
    M          = result["M"]
    new_ghg    = result["new_ghg"]
    new_codes  = result["codes_new"]
    names      = result["industry_names"]
    fd_codes   = result["fd_codes"]

    def _display_names():
        return ["분석업체" if c == "분석업체" else names.get(c, "") for c in new_codes]

    new_names = _display_names()

    def _write_matrix(wb, title, codes, dnames, fd_codes, mat, fd,
                      va=None, tax=None, ghg=None):
        ws = wb.create_sheet(title)
        n = len(codes); n_fd = fd.shape[1]; fd_start = 3 + n
        ws.cell(row=5, column=1, value="상품")
        for j, (c, nm) in enumerate(zip(codes, dnames)):
            ws.cell(row=5, column=3 + j, value=c if c != "분석업체" else None)
            ws.cell(row=6, column=3 + j, value=nm)
        for k, fc in enumerate(fd_codes):
            ws.cell(row=5, column=fd_start + k, value=fc)
            ws.cell(row=6, column=fd_start + k, value=FD_NAMES.get(str(fc), str(fc)))
        for i, (c, nm) in enumerate(zip(codes, dnames)):
            r = 7 + i
            ws.cell(row=r, column=1, value=c if c != "분석업체" else None)
            ws.cell(row=r, column=2, value=nm)
            for j in range(n):
                ws.cell(row=r, column=3 + j, value=float(mat[i, j]))
            for k in range(n_fd):
                ws.cell(row=r, column=fd_start + k, value=float(fd[i, k]))
        if va is not None:
            cur = 7 + n
            for v_i, vname in enumerate(VA_NAMES + [TAX_NAME]):
                ws.cell(row=cur, column=2, value=vname)
                for j in range(n):
                    val = va[v_i, j] if v_i < 3 else tax[j]
                    ws.cell(row=cur, column=3 + j, value=float(val))
                cur += 1
            if ghg is not None:
                ws.cell(row=cur, column=1, value="GHG")
                ws.cell(row=cur, column=2, value="직접온실가스배출(GHG)")
                for j in range(n):
                    ws.cell(row=cur, column=3 + j, value=float(ghg[j]))
        ws.column_dimensions["B"].width = 26
        return ws

    def _write_square(wb, title, codes, dnames, arr):
        ws = wb.create_sheet(title)
        n = len(codes)
        for j, (c, nm) in enumerate(zip(codes, dnames)):
            ws.cell(row=5, column=3 + j, value=c if c != "분석업체" else None)
            ws.cell(row=6, column=3 + j, value=nm)
        for i, (c, nm) in enumerate(zip(codes, dnames)):
            r = 7 + i
            ws.cell(row=r, column=1, value=c if c != "분석업체" else None)
            ws.cell(row=r, column=2, value=nm)
            for j in range(n):
                ws.cell(row=r, column=3 + j, value=float(arr[i, j]))
        ws.column_dimensions["B"].width = 26

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # cases 시트
    ws_c = wb_out.create_sheet("cases")
    ws_c["A1"] = "분석업체 정보"; ws_c["A2"] = "기업명"; ws_c["B2"] = result["company_name"]
    ws_c["A3"] = "소속 산업"; ws_c["B3"] = f"{result['company_code']} ({result['industry_name']})"
    ws_c["A4"] = "계산 모드"; ws_c["B4"] = result["case_label"]
    ws_c["A5"] = "매출액(백만원)"; ws_c["B5"] = result["sales"]
    ws_c["A7"] = "원가 배분 결과"

    # 헤더
    ws_c["A8"] = "산업코드"
    ws_c["B8"] = "산업명"
    ws_c["C8"] = "중간투입액 (백만원)"
    ws_c["D8"] = "비중 (%)"

    # L은 원본 codes 기준 (분석업체 삽입 전), codes_new에서 분석업체 제외한 순서와 동일
    L_vec   = result["L"]      # np.ndarray, shape (n,)
    L_va    = result["L_va"]   # np.ndarray, shape (3,)
    orig_codes = [c for c in result["codes_new"] if c != "분석업체"]
    orig_names = result["industry_names"]
    sales_val  = result["sales"]

    row = 9
    for i, c in enumerate(orig_codes):
        nm  = orig_names.get(c, c)
        val = float(L_vec[i])
        pct = val / sales_val * 100 if sales_val > 0 else 0.0
        ws_c.cell(row=row, column=1, value=c)
        ws_c.cell(row=row, column=2, value=nm)
        ws_c.cell(row=row, column=3, value=round(val, 3))
        ws_c.cell(row=row, column=4, value=round(pct, 4))
        row += 1

    # 중간투입 소계
    total_inter = float(L_vec.sum())
    ws_c.cell(row=row, column=2, value="중간투입 소계")
    ws_c.cell(row=row, column=3, value=round(total_inter, 3))
    ws_c.cell(row=row, column=4, value=round(total_inter / sales_val * 100, 4) if sales_val else 0.0)
    row += 1

    # 부가가치 3항목
    for v_i, vname in enumerate(VA_NAMES):
        val = float(L_va[v_i])
        pct = val / sales_val * 100 if sales_val > 0 else 0.0
        ws_c.cell(row=row, column=2, value=vname)
        ws_c.cell(row=row, column=3, value=round(val, 3))
        ws_c.cell(row=row, column=4, value=round(pct, 4))
        row += 1

    # 부가가치 소계
    total_va = float(L_va.sum())
    ws_c.cell(row=row, column=2, value="부가가치 소계")
    ws_c.cell(row=row, column=3, value=round(total_va, 3))
    ws_c.cell(row=row, column=4, value=round(total_va / sales_val * 100, 4) if sales_val else 0.0)
    row += 1

    # 합계
    ws_c.cell(row=row, column=2, value="합계 (매출액)")
    ws_c.cell(row=row, column=3, value=round(total_inter + total_va, 3))
    ws_c.cell(row=row, column=4, value=round((total_inter + total_va) / sales_val * 100, 4) if sales_val else 0.0)

    ws_c.column_dimensions["A"].width = 10
    ws_c.column_dimensions["B"].width = 28
    ws_c.column_dimensions["C"].width = 22
    ws_c.column_dimensions["D"].width = 14

    # 요약 시트
    ws_s = wb_out.create_sheet("요약")
    ws_s["A1"] = "분석업체명"; ws_s["B1"] = result["company_name"]
    ws_s["A2"] = "소속 산업"; ws_s["B2"] = f"{result['company_code']} ({result['industry_name']})"
    ws_s["A3"] = "매출액(백만원)"; ws_s["B3"] = result["sales"]
    ws_s["A4"] = "계산 모드"; ws_s["B4"] = result["case_label"]
    ws_s["A6"] = "Scope1 (직접배출, tCO2eq.)"; ws_s["B6"] = result["scope1"]
    ws_s["A7"] = "Scope2 (전력 간접배출, tCO2eq.)"; ws_s["B7"] = result["scope2"]
    ws_s["A8"] = "Scope3 (기타 간접배출, tCO2eq.)"; ws_s["B8"] = result["scope3"]
    ws_s["A9"] = "총 탄소배출량 (tCO2eq.)"; ws_s["B9"] = result["total_emission"]
    ws_s.column_dimensions["A"].width = 30

    # 거래표·계수 시트
    _write_matrix(wb_out, "A표_총거래표(생산자)_new", new_codes, new_names, fd_codes,
                  A_new["mat"], A_new["fd"], A_new["va"], A_new["tax"], ghg=new_ghg)
    _write_matrix(wb_out, "국산거래표_new", new_codes, new_names, fd_codes,
                  D_new["mat"], D_new["fd"])
    _write_square(wb_out, "Ad", new_codes, new_names, Ad)
    _write_square(wb_out, "Lf", new_codes, new_names, Lf)
    _write_square(wb_out, "M", new_codes, new_names, M)

    safe = "".join(ch for ch in result["company_name"] if ch.isalnum() or ch in " _-").strip() or "기업"
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out  = output_dir / f"{safe}_{result['company_code']}_탄소배출_{ts}.xlsx"
    wb_out.save(out)
    print(f"💾 결과 저장: {out}")
    return out
